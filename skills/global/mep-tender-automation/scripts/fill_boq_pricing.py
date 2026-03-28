# -*- coding: utf-8 -*-
"""
Fill BoQ pricing for Core5 HP II MEP Tender — FIXED VERSION
Corrected column mappings based on debug analysis:
- B1: C1=Item#, C3=Unit, C5=Rate, C6=Amount
- B2: C11=Mat, C12=Lab, C13=Total(rate), C14=Amount, C15=Contractor Amt
- B3: C12=Mat, C13=Lab, C14=Total(rate), C15=Amount, C16=Contractor Amt
- B4: Same as B2
"""
import openpyxl
import sys, io, os, shutil, re
from copy import copy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\A. COMMERCIAL PART\3. Priced Bill of Quantity & Final Summary\Submit TT-18112025-C5HP2-MEP-BOQ-R3  to Issue Tender.xlsx"
DST = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\A. COMMERCIAL PART\3. Priced Bill of Quantity & Final Summary\TT-C5HP2-MEP-BOQ-R3-PRICED.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

# ============================================================
# PRICE DATABASE
# ============================================================

PRELIM_RATES = {
    1: 0, 2: 0, 3: 50_000_000, 4: 35_000_000, 5: 0,
    6: 15_000_000, 7: 25_000_000, 8: 45_000_000, 9: 30_000_000,
    10: 20_000_000, 11: 25_000_000, 12: 10_000_000, 13: 15_000_000,
    14: 20_000_000, 15: 0, 16: 0, 17: 0, 18: 15_000_000,
    19: 10_000_000, 20: 0, 21: 30_000_000, 22: 40_000_000,
    23: 15_000_000, 24: 0, 25: 10_000_000, 26: 0, 27: 20_000_000,
    28: 25_000_000, 29: 0, 30: 35_000_000, 31: 25_000_000,
    32: 15_000_000, 34: 40_000_000, 35: 50_000_000, 36: 30_000_000,
    37: 0, 38: 0, 39: 35_000_000, 40: 20_000_000, 41: 10_000_000,
}

# Rate DB: pattern → (material, labour) per unit
RATE_DB = [
    # ===== MV 24KV CABLES =====
    (r"MV cable 24KV.*3C-240mm", 850_000, 120_000),
    (r"MV cable 24KV.*3C-185mm", 650_000, 120_000),
    (r"MV cable 24KV.*3C-150mm", 520_000, 120_000),
    (r"MV cable 24KV.*3C-120mm", 420_000, 120_000),
    (r"MV cable 24KV.*3C-50mm", 220_000, 100_000),
    (r"MV manhole", 8_500_000, 3_500_000),
    (r"HDPE 195/150", 180_000, 55_000),
    (r"Head cable 24KV.*240", 3_200_000, 800_000),
    (r"Head cable 24KV.*185", 2_800_000, 800_000),
    (r"Head cable 24KV.*150", 2_500_000, 800_000),
    (r"Head cable 24KV.*120", 2_200_000, 800_000),
    (r"Head cable 24KV.*50", 1_800_000, 600_000),
    (r"11,130 kVA connection", 350_000_000, 50_000_000),

    # ===== TRANSFORMER STATIONS (sub-items priced individually) =====
    (r"Incoming section LBS", 85_000_000, 12_000_000),
    (r"Incoming section DS.*CB", 95_000_000, 12_000_000),
    (r"Transformer section DS", 120_000_000, 15_000_000),
    (r"Outgoing section LBS", 75_000_000, 10_000_000),
    (r"Transformer 3P.*2000KVA", 680_000_000, 45_000_000),
    (r"Transformer 3P.*1600KVA", 580_000_000, 40_000_000),
    (r"Transformer 3P.*320KVA", 280_000_000, 25_000_000),
    (r"MSB-A1 \(", 250_000_000, 35_000_000),
    (r"MSB-A2 \(", 250_000_000, 35_000_000),
    (r"MSB-A3\.1 \(", 250_000_000, 35_000_000),
    (r"MSB-A3\.2 \(", 220_000_000, 32_000_000),
    (r"MSB-A4\.1 \(", 220_000_000, 32_000_000),
    (r"MSB-A4\.2 \(", 220_000_000, 32_000_000),
    (r"MSB-UT \(", 120_000_000, 18_000_000),
    (r"Cu/XLPE/PVC 3C-120mm.*24kV", 420_000, 80_000),
    (r"Cu/XLPE/PVC 3C-50mm.*24kV", 220_000, 65_000),
    (r"CXV 3x6x1C-300mm", 1_850_000, 180_000),
    (r"CXV 3x5x1C-300mm", 1_600_000, 170_000),
    (r"CXV 3x2x1C 150mm", 850_000, 120_000),
    (r"enclosure shall be", 180_000_000, 25_000_000),
    (r"Oil pie 1m3", 35_000_000, 8_000_000),
    (r"All accessories.*auxiliary", 25_000_000, 8_000_000),
    (r"Connection fees.*EVN", 80_000_000, 15_000_000),

    # ===== GENERATOR =====
    (r"Generator.*300KVA", 650_000_000, 45_000_000),
    (r"Earthing for Generator", 35_000_000, 12_000_000),
    (r"Generator concrete", 15_000_000, 8_000_000),
    (r"Cost for connection.*testing.*commissioning", 25_000_000, 10_000_000),

    # ===== DISTRIBUTION BOARDS (Bill 3 - Infrastructure) =====
    (r"Distribution board MSB-UT.*CAP", 150_000_000, 20_000_000),
    (r"Distribution board DB-GH", 35_000_000, 8_000_000),
    (r"Distribution board DB-FM", 25_000_000, 6_000_000),
    (r"Distribution board DB-WWT", 35_000_000, 8_000_000),
    (r"Distribution board DB-PH", 25_000_000, 6_000_000),
    (r"Distribution board DB-FP", 85_000_000, 12_000_000),
    (r"Distribution board DB-A.*-SM", 45_000_000, 8_000_000),
    (r"Distribution board DB-MD", 45_000_000, 8_000_000),
    # Bill 4 - Office DBs
    (r"Distribution board DB-A.*-UT", 25_000_000, 6_000_000),

    # ===== CABLES - INSIDE FACTORY (Bill 2) =====
    (r"Cable ladder 200x100", 180_000, 45_000),
    (r"Isolator.*2P-20A", 250_000, 80_000),
    (r"Isolator.*3P-20A", 350_000, 80_000),
    (r"Cable for roller.*CV 2x1C 2\.5", 32_000, 18_000),
    (r"Cable for Fire-rated.*CXV/Fr.*4C 2\.5", 55_000, 22_000),
    (r"Cable for exhaust.*CVV.*3C 2\.5", 30_000, 18_000),
    (r"Cable for water heater.*4mm", 45_000, 20_000),
    (r"Cable for dock leveler.*4C 2\.5", 42_000, 20_000),
    (r"Cable for fire alarm.*CXV/Fr.*3C 1\.5", 45_000, 20_000),
    (r"PVC conduit D20", 12_000, 8_000),
    (r"PVC conduit D32", 18_000, 10_000),

    # ===== CABLES - LV (Bill 3 Infrastructure) =====
    (r"HDPE 190/150", 160_000, 50_000),
    (r"HDPE 130/110", 120_000, 45_000),
    (r"HDPE 65/50 underground", 65_000, 35_000),
    (r"HDPE D65/50", 65_000, 35_000),
    (r"HDPE D130/110", 120_000, 45_000),
    (r"HDPE D30", 35_000, 25_000),
    (r"HDPE underground.*D50/40", 55_000, 30_000),
    (r"LV manhole", 5_500_000, 2_500_000),

    # Cables from MSB/DB
    (r"From MSB.*to DB-FM.*10mm", 120_000, 30_000),
    (r"From MSB.*to DB-PH.*35mm", 280_000, 45_000),
    (r"From MSB.*to DB-WWT.*50mm", 380_000, 55_000),
    (r"From GEN to MSB", 1_850_000, 180_000),
    (r"From MSB-UT to DB.*95mm", 650_000, 85_000),
    (r"From MSB-UT to DB.*185mm", 1_200_000, 130_000),
    (r"From MSB-UT to DB.*150mm", 950_000, 110_000),
    (r"From MSB-UT to MSB.*50mm", 380_000, 55_000),
    (r"From MSB-A4.*DB-GH.*50mm", 380_000, 55_000),
    (r"From MSB-UT.*GH2.*35mm", 280_000, 45_000),
    (r"From MSB-A4.*DB-FP.*120", 750_000, 95_000),
    (r"From DB.*SM to DB.*2x1C 95mm", 650_000, 85_000),
    (r"From DB.*SM to DB.*1C 95mm", 450_000, 65_000),
    (r"From DB.*SM to DB.*1C 50mm", 280_000, 45_000),
    (r"From DB.*SM to DB.*2x1C 185mm", 1_200_000, 130_000),
    (r"From DB.*SM to DB.*2x1C 120mm", 750_000, 95_000),
    (r"From DB.*SM to DB.*1C 120mm", 520_000, 75_000),
    (r"From DB.*SM to DB.*1C 70mm", 380_000, 55_000),

    # Cable from DB to DB (Bill 4)
    (r"Cable from DB-Axx.*CVV.*6mm", 85_000, 25_000),

    # ===== EARTHING =====
    (r"Main earthing bar", 2_500_000, 800_000),
    (r"Bare cable 70mm", 220_000, 35_000),
    (r"Bare cable 50mm", 160_000, 30_000),
    (r"Bare cable 6mm", 25_000, 12_000),
    (r"Copper rod .*16mm.*2.*4m", 850_000, 350_000),
    (r"Earth pit", 1_200_000, 450_000),
    (r"Test box\b", 650_000, 200_000),
    (r"necessary accessories.*auxiliary", 15_000_000, 5_000_000),
    (r"Testing and commissioning.*cert", 25_000_000, 8_000_000),

    # ===== ELV / CCTV =====
    (r"ELV manhole 600", 3_500_000, 1_500_000),
    (r"ELV manhole 1700", 8_500_000, 3_500_000),
    (r"Cable trunking 100x100", 85_000, 25_000),
    (r"PABX 3CO", 12_000_000, 2_000_000),
    (r"ODF\b", 3_500_000, 800_000),
    (r"ODF 4port", 2_500_000, 600_000),
    (r"Lan data socket", 180_000, 80_000),
    (r"Telephone socket", 150_000, 80_000),
    (r"UTP Cat6", 12_000, 6_000),
    (r"Weather proof.*bullet.*2MP", 3_500_000, 500_000),
    (r"Weather proof.*bullet.*4MP", 5_500_000, 500_000),
    (r"IP digital video", 35_000_000, 3_000_000),
    (r"40.*monitor", 8_500_000, 500_000),
    (r"UPS 5KVA", 25_000_000, 2_000_000),
    (r"Workstation computer", 18_000_000, 500_000),
    (r"Fiber optic 12FO", 35_000, 15_000),
    (r"Fiber optic 4FO", 22_000, 12_000),
    (r"Driver 20TB", 12_000_000, 500_000),
    (r"Rack cabinet 42U", 45_000_000, 5_000_000),
    (r"Rack.*10U.*24 port", 18_000_000, 2_000_000),
    (r"Rack.*10U.*16 port", 15_000_000, 2_000_000),
    (r"Rack.*10U.*12 port", 14_000_000, 2_000_000),
    (r"Rack.*10U.*8 port", 12_000_000, 2_000_000),
    (r"ELV box.*ODF", 5_500_000, 1_200_000),
    (r"Camera pole", 3_500_000, 1_200_000),

    # ===== OUTDOOR LIGHTING =====
    (r"Streetlight.*LED.*100W.*pole", 8_500_000, 2_500_000),
    (r"Floodlight.*LED.*100W", 3_200_000, 800_000),
    (r"Power cable CXV.*4C 2\.5", 42_000, 20_000),

    # ===== CABLES - Guards/Utility (Bill 3 Section) =====
    (r"Cable for lighting.*1\.5mm", 25_000, 15_000),
    (r"Cable for logo.*1\.5mm", 28_000, 15_000),
    (r"Cable for emergency.*3C 2\.5", 48_000, 20_000),
    (r"Cable for socket.*2\.5mm", 32_000, 18_000),
    (r"Cable for AC CVV.*2\.5", 30_000, 18_000),
    (r"Cable for gate motor", 42_000, 20_000),
    (r"Cable for rack and fire.*6mm", 65_000, 22_000),
    (r"Cable for rack cabinet.*1\.5", 25_000, 15_000),
    (r"Cable for PA.*4mm", 55_000, 20_000),
    (r"Cable for exhaust.*2\.5mm", 30_000, 18_000),
    (r"Cable for DB-RO.*4C 4mm", 55_000, 22_000),
    (r"Cable for DB-DWP.*25mm", 180_000, 35_000),
    (r"Cable Cu/PVC 1C-1\.5mm", 8_000, 5_000),
    (r"Cable Cu/PVC 1C-2\.5mm", 12_000, 5_500),
    (r"Cable Cu/PVC/PVC 2C-2\.5", 22_000, 12_000),
    (r"Cable Cu/PVC/PVC 3C-1\.5", 22_000, 12_000),
    (r"Cable Cu/PVC/PVC 3C-2\.5", 28_000, 14_000),

    # ===== INDOOR LIGHTING (Bill 3 Guard+UT) =====
    (r"Recessed Downlight.*LED.*9W", 280_000, 65_000),
    (r"LED lamp TL8-9W", 180_000, 55_000),
    (r"LED lamp 2xTL8-18W.*batten", 380_000, 65_000),
    (r"LED lamp 2xTL8-18W.*Weather", 480_000, 70_000),
    (r"LED lamp TL8-18W.*batten.*wall", 250_000, 60_000),
    (r"LED lamp TL8-18W.*Weather", 380_000, 65_000),
    (r"LED lamp panel 40W", 650_000, 80_000),
    (r"Hibay.*LED 100W", 2_800_000, 350_000),

    # ===== SWITCHES / SENSORS =====
    (r"1 way single switch", 65_000, 25_000),
    (r"1 way double switch", 85_000, 30_000),
    (r"1 way three switch", 105_000, 35_000),
    (r"1 way third switch", 105_000, 35_000),
    (r"1 way four switch", 125_000, 40_000),
    (r"2 way.*switch", 85_000, 30_000),
    (r"1 way switch", 65_000, 25_000),
    (r"PIR sensor.*ceiling", 450_000, 80_000),
    (r"PIR sensor.*wall", 380_000, 80_000),
    (r"Grounding duplex.*receptacle", 120_000, 35_000),
    (r"Double socket 16A", 120_000, 35_000),
    (r"Celling fan wall", 180_000, 50_000),

    # ===== PLUMBING - WATER SUPPLY =====
    (r"Underground HDPE.*D100", 95_000, 35_000),
    (r"Underground HDPE.*D8\b", 75_000, 30_000),
    (r"Underground HDPE.*D65", 60_000, 28_000),
    (r"Underground HDPE.*D50", 48_000, 25_000),
    (r"Underground HDPE.*D40", 38_000, 22_000),
    (r"Underground HDPE.*D25", 28_000, 18_000),
    (r"HPDE D50", 48_000, 25_000),
    (r"HPDE D40", 38_000, 22_000),
    (r"HPDE D32", 32_000, 20_000),
    (r"HPDE D25", 28_000, 18_000),
    (r"HPDE D20", 22_000, 15_000),
    (r"HDPE DN40 pipe", 38_000, 22_000),
    (r"PPR D50", 65_000, 22_000),
    (r"PPR D40", 48_000, 18_000),
    (r"PPR D32", 38_000, 15_000),
    (r"PPR D25", 28_000, 12_000),
    (r"PPR DN25", 28_000, 12_000),
    (r"PPR D20", 22_000, 10_000),
    (r"PPR DN20", 22_000, 10_000),
    (r"PPR D15", 18_000, 8_000),
    (r"PPR DN15", 18_000, 8_000),
    (r"PPR valve", 75_000, 22_000),
    (r"Float valve", 850_000, 150_000),
    (r"Check valve D100", 1_200_000, 200_000),
    (r"Check valve DN100", 1_200_000, 200_000),
    (r"Check valve D50", 650_000, 120_000),
    (r"Check valve DN50", 650_000, 120_000),
    (r"Check valve D40", 450_000, 100_000),
    (r"Check valve DN40", 450_000, 100_000),
    (r"Check valve DN80", 950_000, 180_000),
    (r"Gate valve D100", 1_500_000, 250_000),
    (r"Gate valve DN100", 1_500_000, 250_000),
    (r"Gate valve D65", 950_000, 200_000),
    (r"Gate valve DN80", 1_200_000, 220_000),
    (r"Gate valve DN50", 750_000, 150_000),
    (r"Gate valve DN40", 550_000, 120_000),
    (r"Valve D40", 450_000, 100_000),
    (r"Valve D25", 350_000, 80_000),
    (r"Valve D20", 280_000, 70_000),
    (r"Water meter.*D40$", 2_500_000, 350_000),
    (r"Water meter.*D50", 3_200_000, 400_000),
    (r"Water meter.*D80", 4_500_000, 500_000),
    (r"Main water meter", 8_500_000, 800_000),
    (r"Water supply.*booster pump", 85_000_000, 12_000_000),
    (r"Control panel", 15_000_000, 3_000_000),
    (r"Power supply cable.*10mm", 120_000, 30_000),
    (r"Power supply cable.*2\.5", 30_000, 15_000),
    (r"Control cable for pump", 8_000_000, 2_000_000),
    (r"Cable trays.*conduits.*pump", 12_000_000, 3_000_000),
    (r"Conduits for power.*pump", 5_000_000, 1_500_000),
    (r"Water tap for floor", 350_000, 80_000),
    (r"Water tap\b", 350_000, 80_000),
    (r"Tap D15", 250_000, 60_000),
    (r"Y Strainer DN40", 350_000, 80_000),
    (r"Y Strainer DN50", 450_000, 100_000),
    (r"Y Strainer DN80", 650_000, 150_000),
    (r"Y Strainer DN100", 850_000, 180_000),
    (r"Cast iron valve cover", 1_800_000, 300_000),
    (r"Concrete water meter manhole", 3_500_000, 1_500_000),
    (r"Brick gate valve manhole", 2_800_000, 1_200_000),
    (r"Foundation for Pumps", 12_000_000, 5_000_000),
    (r"Foot valve", 1_500_000, 250_000),
    (r"Level Swich", 1_200_000, 200_000),
    (r"Pressure switch", 1_500_000, 250_000),
    (r"Pressure gauge", 850_000, 150_000),
    (r"Pressure sensor", 2_500_000, 300_000),
    (r"Flexible connection", 450_000, 100_000),
    (r"RO filter booster pump", 25_000_000, 3_000_000),
    (r"RO system", 120_000_000, 15_000_000),
    (r"G\.I pipe DN125", 180_000, 45_000),
    (r"G\.I pipe DN100", 150_000, 40_000),
    (r"G\.I pipe DN80", 120_000, 35_000),
    (r"G\.I pipe DN65", 95_000, 30_000),
    (r"G\.I pipe DN40", 65_000, 25_000),
    (r"Fitting for above", 5_000_000, 1_500_000),
    (r"Support above", 4_000_000, 1_200_000),
    (r"Painting work", 3_000_000, 1_500_000),
    (r"Water meter.*factory.*D40", 2_500_000, 350_000),
    (r"Water meter.*IR.*D50", 3_200_000, 400_000),
    (r"Water meter.*pump.*D80", 4_500_000, 500_000),

    # ===== DRAINAGE =====
    (r"uPVC D100 pipe", 65_000, 22_000),
    (r"uPVC D80 pipe", 52_000, 20_000),
    (r"uPVC D75 pipe", 48_000, 18_000),
    (r"uPVC D50 pipe", 38_000, 15_000),
    (r"uPVC D48 pipe", 36_000, 15_000),
    (r"uPVC D42 pipe", 32_000, 14_000),
    (r"uPVC D40 pipe", 30_000, 14_000),
    (r"uPVC pipe D80", 52_000, 20_000),
    (r"uPVC pipe D100", 65_000, 22_000),
    (r"uPVC pipe D125", 85_000, 28_000),
    (r"uPVC pipe D150", 110_000, 32_000),
    (r"uPVC pipe D200", 165_000, 40_000),
    (r"uPVC pipe D250", 220_000, 48_000),
    (r"uPVC Fitting", 5_000_000, 1_500_000),
    (r"Floor drainage", 280_000, 80_000),

    # ===== SANITARY =====
    (r"Lavatory Pan", 2_800_000, 350_000),
    (r"Paper holder", 280_000, 50_000),
    (r"Bidet sprayer", 380_000, 60_000),
    (r"Lavatory aucet", 1_200_000, 150_000),
    (r"^Lavatory$", 1_800_000, 250_000),
    (r"Lavatory\b", 1_800_000, 250_000),
    (r"Wall-mounted mirror", 450_000, 80_000),
    (r"Uniral", 2_500_000, 350_000),
    (r"Shower\b", 2_200_000, 300_000),
    (r"Water heater 30L", 4_500_000, 350_000),
    (r"Water tank 2\.5m3", 8_500_000, 1_200_000),
    (r"Soap dispenser", 250_000, 50_000),
    (r"Clean out DN100", 180_000, 60_000),
    (r"Thông tắc.*DN80", 150_000, 55_000),
    (r"Thông tắc.*DN100", 180_000, 60_000),
    (r"đựng xà phòng", 250_000, 50_000),

    # ===== STORM WATER =====
    (r"Siphonic roof", 3_500_000, 800_000),
    (r"Coventional oulet D200", 1_200_000, 350_000),
    (r"Coventional oulet D125", 850_000, 250_000),
    (r"Coventional oulet D100", 650_000, 200_000),
    (r"Coventional oulet D80", 450_000, 150_000),
    (r"HDPE pipe D75", 55_000, 22_000),
    (r"HDPE pipe D90", 72_000, 25_000),
    (r"HDPE pipe D110", 95_000, 30_000),
    (r"HDPE pipe D125", 120_000, 35_000),
    (r"HDPE pipe D160", 160_000, 42_000),
    (r"HDPE pipe D200", 220_000, 50_000),
    (r"HDPE fittings", 8_000_000, 2_500_000),

    # ===== LIGHTNING =====
    (r"Lightning arrester.*R107", 25_000_000, 5_000_000),
    (r"GI pipe 5m.*foundation", 8_500_000, 3_500_000),
    (r"Steel wire", 35_000, 12_000),
    (r"Copper cable 70mm", 220_000, 35_000),
    (r"Earth rod.*16mm", 850_000, 350_000),
    (r"Test box.*metal", 650_000, 200_000),
    (r"Concrete inspection", 1_500_000, 600_000),
    (r"5m deep grounding", 5_500_000, 2_500_000),
    (r"Lightning event", 3_500_000, 500_000),
    (r"Testing and Commissioning\b", 8_000_000, 3_000_000),

    # ===== ACMV =====
    (r"Air conditioner.*17", 15_000_000, 2_500_000),
    (r"Air conditioner.*12", 10_000_000, 2_000_000),
    (r"Wall mounted exhaust.*255", 2_500_000, 500_000),
    (r"Wall mounted exhaust.*546", 3_500_000, 600_000),
    (r"Wall mounted exhaust.*90", 1_500_000, 400_000),
    (r"Ceiling mounted.*330", 3_200_000, 600_000),
    (r"Inline fan 4200", 8_500_000, 1_500_000),
    (r"Inline fan 2000", 5_500_000, 1_200_000),
    (r"Inline fan 1300", 4_500_000, 1_000_000),
    (r"Inline fan 1200", 4_200_000, 950_000),
    (r"Inline fan 770", 3_500_000, 800_000),
    (r"Exhaust fan 1300", 4_500_000, 1_000_000),
    (r"Thermostat", 2_500_000, 350_000),
    (r"Copper piping.*12\.7", 280_000, 85_000),
    (r"Copper piping.*9\.5", 220_000, 75_000),
    (r"uPVC DN21", 35_000, 12_000),
    (r"Air grill 1000x400", 850_000, 180_000),
    (r"Air grill 250x250", 380_000, 120_000),
    (r"Air grill 400x400", 550_000, 150_000),
    (r"Air louver 600x300", 650_000, 150_000),
    (r"Air louver 500x300", 550_000, 130_000),
    (r"Air louver 300x300", 380_000, 120_000),
    (r"VCD D100", 850_000, 200_000),
    (r"Louver 900x600", 1_200_000, 250_000),
    (r"Louver 500x300", 650_000, 180_000),
    (r"Louver 300x300", 450_000, 150_000),
    (r"Motorize damper 300x500", 3_500_000, 500_000),
    (r"Motorize damper 900x600", 5_500_000, 700_000),
    (r"Galvanized duct.*0\.58", 250_000, 85_000),
    (r"Galvanized duct.*0\.75", 320_000, 95_000),
    (r"Flexible duct D100", 120_000, 35_000),
    (r"Ventilation pipe.*D114", 55_000, 20_000),
    (r"Drainage pipe with insulation", 45_000, 15_000),
    (r"Drainage pipe without", 35_000, 12_000),

    # ===== OTHER WORKS (Bill 3 section E) =====
    (r"3rd Testing.*Commiss", 35_000_000, 10_000_000),
    (r"Excavation.*backfill.*HV", 25_000_000, 15_000_000),
    (r"Concrete duct bank.*HV", 45_000_000, 20_000_000),
    (r"Excavation.*backfill.*LV", 30_000_000, 18_000_000),
    (r"Concrete duct bank.*LV", 55_000_000, 25_000_000),
    (r"Cable ladder.*Subtat", 12_000_000, 3_000_000),
    (r"Excavation.*Backfilling", 150, 100),

    # ===== Submersible pumps =====
    (r"Submersible wastewater", 18_000_000, 3_000_000),
    (r"Submersible weighing", 22_000_000, 3_500_000),
]


def find_rate(desc):
    """Find matching rate from ordered list"""
    for pattern, mat, lab in RATE_DB:
        if re.search(pattern, desc, re.IGNORECASE):
            return (mat, lab)
    return None


# ============================================================
# BILL 1 - PRELIM
# ============================================================
ws1 = wb['B1_PRELIM']
filled_b1 = 0
for row in range(7, ws1.max_row + 1):
    c1 = ws1.cell(row=row, column=1).value
    if c1 is None:
        continue
    try:
        item_no = int(str(c1).strip())
    except ValueError:
        continue
    if item_no in PRELIM_RATES:
        rate = PRELIM_RATES[item_no]
        c3 = ws1.cell(row=row, column=3).value
        if c3 and str(c3).strip() == 'Ls':
            ws1.cell(row=row, column=5, value=rate)
            ws1.cell(row=row, column=6, value=rate)  # Qty=1
            if rate > 0:
                filled_b1 += 1

print(f"Bill 1 PRELIM: filled {filled_b1} items")


# ============================================================
# HELPER: Fill a bill sheet
# ============================================================
def fill_bill_sheet(sheet_name, mat_col, lab_col, total_rate_col, amt_col):
    """
    mat_col: Material rate column
    lab_col: Labour rate column
    total_rate_col: Total rate column
    amt_col: Amount column (qty × total rate)
    """
    ws = wb[sheet_name]
    filled = 0
    unfilled = []
    
    for row in range(10, ws.max_row + 1):
        desc_val = ws.cell(row=row, column=3).value
        unit_val = ws.cell(row=row, column=4).value
        
        if desc_val is None:
            continue
        if unit_val is None:
            continue
        
        unit_str = str(unit_val).strip().lower()
        if unit_str in ['', 'note', 'ghi chú']:
            continue
        
        desc = str(desc_val)
        
        # Skip already priced (non-zero material rate)
        existing = ws.cell(row=row, column=mat_col).value
        if existing and existing not in [0, '0', None, '']:
            try:
                if float(existing) > 0:
                    continue
            except:
                pass
        
        rates = find_rate(desc)
        if rates:
            mat_r, lab_r = rates
            total_r = mat_r + lab_r
            
            # Get qty: try contractor qty column, then total column
            qty = None
            # For Bill 2,4: contractor qty = C10, for Bill 3: contractor qty = C11
            if sheet_name == '3_INFRAS+UTILITI AREAS':
                qty = ws.cell(row=row, column=11).value
                if not qty or qty == 0:
                    qty = ws.cell(row=row, column=10).value
            else:
                qty = ws.cell(row=row, column=10).value
                if not qty or qty == 0:
                    qty = ws.cell(row=row, column=9).value
            
            if not qty or qty == 0:
                # Try individual area sums
                for qc in [5, 6, 7, 8, 9]:
                    v = ws.cell(row=row, column=qc).value
                    if v and v != 0:
                        qty = v
                        break
            
            if not qty:
                qty = 1
            try:
                qty = float(qty)
            except:
                qty = 1
            
            ws.cell(row=row, column=mat_col, value=mat_r)
            ws.cell(row=row, column=lab_col, value=lab_r)
            ws.cell(row=row, column=total_rate_col, value=total_r)
            ws.cell(row=row, column=amt_col, value=round(total_r * qty))
            filled += 1
        else:
            if unit_str not in ['', 'note', 'ghi chú']:
                unfilled.append(f"R{row}: [{unit_str}] {desc[:60]}")
    
    print(f"{sheet_name}: filled {filled} items, {len(unfilled)} unfilled")
    for u in unfilled[:5]:
        print(f"   ? {u}")
    if len(unfilled) > 5:
        print(f"   ... +{len(unfilled)-5} more")
    return filled


# ============================================================
# FILL BILLS 2, 3, 4
# ============================================================
# Bill 2: mat=C11, lab=C12, total_rate=C13, amount=C14
fill_bill_sheet('2_INSIDE FACTORY', 11, 12, 13, 14)

# Bill 3: mat=C12, lab=C13, total_rate=C14, amount=C15
fill_bill_sheet('3_INFRAS+UTILITI AREAS', 12, 13, 14, 15)

# Bill 4: mat=C11, lab=C12, total_rate=C13, amount=C14
fill_bill_sheet('4_FACTORY OFFICES', 11, 12, 13, 14)


# ============================================================
# UPDATE SUMMARY
# ============================================================
def calc_total(sheet_name, amt_col):
    ws = wb[sheet_name]
    total = 0
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=amt_col).value
        if v and isinstance(v, (int, float)) and v > 0:
            total += v
    return total

b1 = sum(r for r in PRELIM_RATES.values() if r > 0)
b2 = calc_total('2_INSIDE FACTORY', 14)
b3 = calc_total('3_INFRAS+UTILITI AREAS', 15)
b4 = calc_total('4_FACTORY OFFICES', 14)
gt = b1 + b2 + b3 + b4

ws_s = wb['SUMMARY']
ws_s.cell(row=10, column=5, value=b1)
ws_s.cell(row=12, column=5, value=b2)
ws_s.cell(row=14, column=5, value=b3)
ws_s.cell(row=16, column=5, value=b4)
ws_s.cell(row=18, column=5, value=gt)

print(f"\n{'='*50}")
print(f"SUMMARY")
print(f"{'='*50}")
print(f"B1 Prelim:       {b1:>18,.0f} VND")
print(f"B2 Inside:       {b2:>18,.0f} VND")
print(f"B3 Infra:        {b3:>18,.0f} VND")
print(f"B4 Office:       {b4:>18,.0f} VND")
print(f"{'─'*50}")
print(f"TOTAL (ex VAT):  {gt:>18,.0f} VND")
print(f"VAT 10%:         {gt*0.1:>18,.0f} VND")
print(f"GRAND TOTAL:     {gt*1.1:>18,.0f} VND")

wb.save(DST)
wb.close()
print(f"\nSaved: {os.path.basename(DST)} ({os.path.getsize(DST)/1024:.1f} KB)")
