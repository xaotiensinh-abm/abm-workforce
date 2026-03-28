# -*- coding: utf-8 -*-
"""Fill empty C4 (Tenderer's Offers) cells in all 3 Technical Data Excel files"""
import openpyxl, os, shutil, re, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\B. TECHNICAL PART\9. Technical Data"

files = {
    "TECHNICAL DATA_ACMV_Ver1.xlsx": "ACVM",
    "TECHNICAL DATA_PLUMBING_Ver2.xlsx": "PD FF",
    "TECHNICAL DATA _ELEC_Ver1.xlsx": "ELEC",
}

# Process each file
for fn, sheet_name in files.items():
    src = os.path.join(BASE, fn)
    dst = os.path.join(BASE, fn.replace(".xlsx", "-FILLED.xlsx"))
    shutil.copy2(src, dst)
    
    wb = openpyxl.load_workbook(dst)
    ws = wb[sheet_name]
    
    filled = 0
    total_empty = 0
    
    for row in range(6, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value  # NO
        c2 = ws.cell(row=row, column=2).value  # Description
        c3 = ws.cell(row=row, column=3).value  # Specified
        c4 = ws.cell(row=row, column=4).value  # Tenderer's Offers
        
        # Skip section headers, empty rows
        if c2 is None:
            continue
        
        desc = str(c2).replace('\n', ' ').strip()
        
        # Only fill if C3 has value (specified) but C4 is empty
        if c3 is not None and (c4 is None or str(c4).strip() == ''):
            spec = str(c3).replace('\n', ' ').strip()
            total_empty += 1
            
            # Default strategy: comply with specification
            offer = None
            
            # Smart matching for common pattern
            if 'refer' in spec.lower() or 'xem' in spec.lower() or 'drawing' in spec.lower():
                offer = "Comply tender drawings\nTheo bản vẽ thiết kế"
            elif 'catalog' in spec.lower():
                offer = "Follow manufactures\nTheo thông số của nhà sản xuất"
            elif 'asia' in spec.lower():
                offer = "Asia"
            elif 'viet' in spec.lower():
                offer = "Viet Nam"
            elif spec.lower() in ['yes', 'có', 'có|yes']:
                offer = "Có|Yes"
            elif 'iso' in spec.lower() or 'iec' in spec.lower() or 'tcvn' in spec.lower():
                offer = "Đáp ứng / Comply"
            elif 'ip' in spec.lower() and any(c.isdigit() for c in spec):
                offer = spec  # Keep same IP rating
            elif any(x in spec.lower() for x in ['mm', 'kv', 'kva', 'rpm', 'pa', 'db', 'hz', 'bar']):
                offer = spec  # Technical values: comply
            elif 'copper' in spec.lower() or 'đồng' in spec.lower():
                offer = spec
            elif 'epoxy' in spec.lower() or 'galvanized' in spec.lower():
                offer = spec
            elif len(spec) < 80:
                offer = spec  # Short specs: comply as-is
            else:
                offer = "Đáp ứng / Comply"
            
            if offer:
                ws.cell(row=row, column=4, value=offer)
                filled += 1
    
    wb.save(dst)
    wb.close()
    print(f"{fn}: filled {filled}/{total_empty} empty C4 cells -> {os.path.basename(dst)}")

print("\nDone!")
