# -*- coding: utf-8 -*-
"""Extract FULL BoQ data (all rows) from all sheets"""
import openpyxl, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\A. COMMERCIAL PART\3. Priced Bill of Quantity & Final Summary\Submit TT-18112025-C5HP2-MEP-BOQ-R3  to Issue Tender.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

out = open(r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\boq_full_extract.txt", "w", encoding="utf-8")

for sn in wb.sheetnames:
    ws = wb[sn]
    out.write(f"\n{'='*70}\n")
    out.write(f"SHEET: {sn}\n")
    out.write(f"{'='*70}\n")
    for i, row in enumerate(ws.iter_rows(max_col=14, values_only=True)):
        vals = list(row)
        if not any(v for v in vals if v is not None):
            continue
        parts = []
        for j, v in enumerate(vals):
            if v is not None:
                s = str(v)[:70]
                parts.append(f"C{j+1}={s}")
        if parts:
            out.write(f"  R{i+1}: {' | '.join(parts)}\n")

wb.close()
out.close()
print("Done!")
