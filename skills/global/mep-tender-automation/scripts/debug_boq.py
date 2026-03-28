# -*- coding: utf-8 -*-
"""Debug B1 item matching"""
import openpyxl, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
path = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\A. COMMERCIAL PART\3. Priced Bill of Quantity & Final Summary\Submit TT-18112025-C5HP2-MEP-BOQ-R3  to Issue Tender.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb['B1_PRELIM']

out = open(r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\debug_b1.txt", "w", encoding="utf-8")
for row in range(15, 60):
    c1 = ws.cell(row=row, column=1).value
    c3 = ws.cell(row=row, column=3).value
    c4 = ws.cell(row=row, column=4).value
    c5 = ws.cell(row=row, column=5).value
    c6 = ws.cell(row=row, column=6).value
    out.write(f"R{row}: C1={repr(c1)} C3={repr(c3)} C4={repr(c4)} C5={repr(c5)} C6={repr(c6)}\n")

# Also check B3 unfilled MV cables
ws3 = wb['3_INFRAS+UTILITI AREAS']
out.write("\n=== B3 MV cable rows ===\n")
for row in [21, 26, 31, 36, 41]:
    c3 = ws3.cell(row=row, column=3).value
    out.write(f"R{row}: {repr(c3)}\n")

wb.close()
out.close()
print("Done!")
