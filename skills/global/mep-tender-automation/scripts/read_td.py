# -*- coding: utf-8 -*-
"""Extract Technical Data structure from all 3 files"""
import openpyxl

base = r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\B. TECHNICAL PART\9. Technical Data"
files = [
    "TECHNICAL DATA_ACMV_Ver1.xlsx",
    "TECHNICAL DATA_PLUMBING_Ver2.xlsx",
    "TECHNICAL DATA _ELEC_Ver1.xlsx",
]

out = open(r"D:\Thang tien\1. TT-CORE5-HP-TENDER DOCCUMENT (1st TIME)\td_extract.txt", "w", encoding="utf-8")

for fn in files:
    path = f"{base}\\{fn}"
    out.write(f"\n{'='*60}\n")
    out.write(f"FILE: {fn}\n")
    out.write(f"{'='*60}\n")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            out.write(f"\n  SHEET: {sn} (max_row={ws.max_row}, max_col={ws.max_column})\n")
            for r in range(1, min(ws.max_row + 1, 80)):
                parts = []
                for c in range(1, min(ws.max_column + 1, 15)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        s = str(v).replace('\n', '|')[:55]
                        parts.append(f"C{c}=[{s}]")
                if parts:
                    out.write(f"    R{r}: {' | '.join(parts)}\n")
        wb.close()
    except Exception as e:
        out.write(f"  ERROR: {e}\n")

out.close()
print("Done!")
